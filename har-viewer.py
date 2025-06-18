import json
import pandas as pd
import os
from datetime import datetime, timedelta
import pytz
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font
from collections import defaultdict

# 读取 HAR 文件
script_dir = os.path.dirname(os.path.abspath(__file__))  # 获取脚本所在目录
file_folder = os.path.join(script_dir, "har")
file_str = "验收测试"
file_path = os.path.join(file_folder, file_str+".har")
save_path = os.path.join(file_folder, file_str+'.xlsx')

with open(file_path, 'r', encoding='utf-8') as f:
    har_data = json.load(f)

# 获取所有的请求条目
entries = har_data['log']['entries']

# 创建一个空的列表来存储 WebSocket 请求
ws_requests = []

# 定义东八区时区
east8 = pytz.timezone('Asia/Shanghai')

# 过滤 WebSocket 请求并提取时间戳和持续时间
for entry in entries:
    request = entry['request']
    url = request['url']
    
    if url.startswith('ws://') or url.startswith('wss://'):
        started_time = entry['startedDateTime']  # 获取请求的开始时间
        duration = entry['time'] / 1000  # 将持续时间从毫秒转换为秒
        # 将时间戳转换为 datetime 对象，并转换为东八区时间
        utc_time = datetime.strptime(started_time, "%Y-%m-%dT%H:%M:%S.%fZ")
        local_time = utc_time.replace(tzinfo=pytz.utc).astimezone(east8)
        # 格式化为 24 小时制
        formatted_time = local_time.strftime("%Y-%m-%d %H:%M:%S")
        # 计算结束时间
        end_time = local_time + timedelta(seconds=duration)
        formatted_end_time = end_time.strftime("%Y-%m-%d %H:%M:%S")
        ws_requests.append({
            'url': url,
            'started_time': formatted_time,
            'end_time': formatted_end_time,
            'duration_seconds': duration
        })

# 按 URL 分组
ws_groups = defaultdict(list)
for request in ws_requests:
    ws_groups[request['url']].append(request)

# 计算相邻 WebSocket 请求的重连间隔
intervals = []
for url, requests in ws_groups.items():
    # 按开始时间排序，确保按时间顺序处理
    requests.sort(key=lambda x: x['started_time'])
    
    # 计算重连间隔
    for i in range(1, len(requests)):
        prev_end_time = datetime.strptime(requests[i - 1]['end_time'], "%Y-%m-%d %H:%M:%S")
        curr_start_time = datetime.strptime(requests[i]['started_time'], "%Y-%m-%d %H:%M:%S")
        time_diff = (curr_start_time - prev_end_time).total_seconds()
        intervals.append({
            'WebSocket URL': url,
            '时间间隔/秒': time_diff,
            '前一次连接结束时间': requests[i - 1]['end_time'],
            '当前连接开始时间': requests[i]['started_time'],
            '持续时间': str(timedelta(seconds=round(requests[i]['duration_seconds'], 3)))
        })

# 将结果转换为 DataFrame
df = pd.DataFrame(intervals)

# 按时间间隔排序
df = df.sort_values(by='当前连接开始时间')

# 添加统计信息
total_reconnects = len(df)
average_interval = df['时间间隔/秒'].mean()
minutes = int(average_interval // 60)
seconds = int(average_interval % 60)
average_interval_formatted = f"{minutes} 分 {seconds} 秒" if minutes > 0 else f"{seconds} 秒"

# 创建统计信息 DataFrame
stats_df = pd.DataFrame({
    '统计项': ['总重连次数', '平均重连间隔'],
    '值': [total_reconnects, average_interval_formatted]
})

# 将统计信息添加到主 DataFrame 中
df = pd.concat([stats_df, df], axis=0)

# 保存到 Excel 文件
with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='数据')
    
    # 获取工作簿和工作表
    workbook = writer.book
    worksheet = writer.sheets['数据']
    
    # 设置表头高亮
    header_fill = PatternFill(start_color="FF71c681", end_color="FF71c681", fill_type="solid")
    header_font = Font(bold=True)  # 加粗字体
    for cell in worksheet[1]:  # 第一行是表头
        cell.fill = header_fill
        cell.font = header_font
    
    # 设置列宽
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # 调整列宽
        adjusted_width = max(adjusted_width, 14) # 确保列宽不小于 14
        worksheet.column_dimensions[column].width = adjusted_width
    
    # 设置单元格对齐方式
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
